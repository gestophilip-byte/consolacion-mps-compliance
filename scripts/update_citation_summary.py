import io
import json
import math
import re
from datetime import datetime, timezone

import requests
from openpyxl import load_workbook

SPREADSHEET_ID = "121UYJVxFOki-djRQN4pq-9fNtEMMfvVi--eWBNClNx0"
EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
OUTPUT_PATH = "citation-summary.json"


def normalize(value):
    return re.sub(r"\s+", " ", str(value or "").strip()).lower()


def number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)

    text = (
        str(value)
        .replace(",", "")
        .replace("₱", "")
        .replace("PHP", "")
        .strip()
    )
    try:
        return float(text)
    except ValueError:
        return None


def find_column(row, phrases):
    for index, value in enumerate(row):
        text = normalize(value)
        if any(phrase in text for phrase in phrases):
            return index
    return None


def is_citation_header(row):
    return find_column(
        row,
        ("total no. of issued citation", "total no of issued citation"),
    ) is not None


def summarize_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    citations = 0.0
    fines = 0.0
    monitoring_blocks = 0
    i = 0

    while i < len(rows):
        header = rows[i]
        citation_col = find_column(
            header,
            ("total no. of issued citation", "total no of issued citation"),
        )
        if citation_col is None:
            i += 1
            continue

        fine_col = find_column(header, ("estimated fine", "estimated fines"))
        if fine_col is None:
            # In many tabs "Estimated fines" sits one row above the main header.
            for lookback in (1, 2):
                if i - lookback < 0:
                    break
                fine_col = find_column(
                    rows[i - lookback],
                    ("estimated fine", "estimated fines"),
                )
                if fine_col is not None:
                    break

        block_citations = 0.0
        block_fines = 0.0
        j = i + 1

        while j < len(rows):
            row = rows[j]
            first_cell = normalize(row[0] if row else None)

            if is_citation_header(row) or first_cell == "total":
                break

            citation_value = number(
                row[citation_col] if citation_col < len(row) else None
            )
            fine_value = number(
                row[fine_col]
                if fine_col is not None and fine_col < len(row)
                else None
            )

            # Sum the underlying personnel rows instead of trusting "Total" cells.
            # Several historical tabs contain mistyped or incomplete total formulas.
            if citation_value is not None:
                block_citations += citation_value
            if fine_value is not None:
                block_fines += fine_value

            j += 1

        citations += block_citations
        fines += block_fines
        monitoring_blocks += 1
        i = max(j, i + 1)

    return citations, fines, monitoring_blocks


def main():
    response = requests.get(EXPORT_URL, timeout=60)
    response.raise_for_status()

    workbook = load_workbook(
        io.BytesIO(response.content),
        data_only=True,
        read_only=True,
    )

    total_citations = 0.0
    overall_fines = 0.0
    populated_sheets = 0
    monitoring_blocks = 0

    for ws in workbook.worksheets:
        citations, fines, blocks = summarize_sheet(ws)
        if citations or fines:
            populated_sheets += 1
        total_citations += citations
        overall_fines += fines
        monitoring_blocks += blocks

    summary = {
        "source": "Citation Monitoring",
        "spreadsheetId": SPREADSHEET_ID,
        "totalCitations": int(round(total_citations)),
        "overallEstimatedFines": round(overall_fines, 2),
        "currency": "PHP",
        "sheetCount": len(workbook.worksheets),
        "populatedSheetCount": populated_sheets,
        "monitoringBlockCount": monitoring_blocks,
        "calculatedAt": datetime.now(timezone.utc)
        .isoformat()
        .replace("+00:00", "Z"),
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as handle:
        json.dump(summary, handle, indent=2)
        handle.write("\n")

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
